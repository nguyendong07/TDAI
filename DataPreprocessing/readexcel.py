import pandas as pd
import os
import xlsxwriter
import csv
from openpyxl import load_workbook

def subtraction(a, b):
    return a - b;

def result(exel_path):
    wb = load_workbook(exel_path)
    sheet = wb.worksheets[0]
    row_count = sheet.max_row
    first_timestamp = sheet.cell(2, 1).value
    last_timestamp = sheet.cell(row_count, 1).value
    filename = exel_path.split("/")[-1]
    timestamp = first_timestamp, last_timestamp
    count = row_count
    sub = subtraction(last_timestamp, first_timestamp)
    return filename, timestamp, count, sub

def csv2xlsx(name_split, path, fullpath):
    read_file = pd.read_csv(fullpath)
    return read_file.to_excel(path +"/" + name_split + ".xlsx", index=None, header=True)
    filtered_files = [file for file in files_in_directory if file.endswith(".csv")]
    for file in filtered_files:
        path_to_file = os.path.join(directory, file)
        os.remove(path_to_file)


def concatname(name):
    txt = name
    split_name = txt.split(".")
    return split_name[0];

def readfoler(path):
    arr = []
    for x in os.walk(path):
        arr.append(x[0])
    return arr

a = []
path="C:/Users/Admin/Desktop/New folder/MobiAct_Dataset_v2.0/Raw Data"
arr = readfoler(path)
del arr[0]
# print(arr)
for path in arr:
   print(path)
   for root, dirs, files in os.walk(path, topdown=False):
        for name in files:
            print(name)
            full_path = path + "/" + name
            # print(full_path)
            with open(full_path) as f:
                for _ in range(15):
                    next(f)
                for line in f:
                    # #
                    if line != "\n":
                        lines = f.readlines()
                        time_start = int([lines[0].split((","))][0][0])
                        time_end = int([lines[-1:][0].split(",")][0][0])
                        time = subtraction(time_end, time_start)
            linecount = 0;
            with open(full_path) as f:
                for _ in range(15):
                    next(f)
                for line in f:
                    if line != "\n":
                        linecount += 1;
                arr = [name, time, time_start, time_end, linecount]
                a.append(arr)

                # print(time,time_start,time_end)
                # print(linecount)
print(a)
with open('C:/Users/Admin/Desktop/test.csv', 'w',newline='\n') as f:
    for i in range(len(a)):
            writer = csv.writer(f)
            writer.writerow(a[i])





            # print(linecount)

# pathsave = "C:/Users/Admin/Desktop/New folder/MobiAct_Dataset_v2.0/Annotated Data"
# arr1 = []
# pathexcel = []
# for path in arr:
#    for root, dirs, files in os.walk(path, topdown=False):
#      for name in files:
#          fullpath = path + "/" + name
#          name_split = concatname(name)
#          csv2xlsx(name_split, path, fullpath)
#
# for path in arr:
#   for root, dirs, files in os.walk(path, topdown=False):
#     for name in files:
#         fullpath = path + "/" + name
#         arr1.append(fullpath)
# with open('C:/Users/Admin/Desktop/test.csv', 'w',newline='\n') as f:
#     for i in range(len(arr1)):
#         if i % 2 != 0:
#             print(arr1[i])
#             a = result(arr1[i])
#             print(type(a))
#             writer = csv.writer(f)
#             writer.writerow(a)
#
#
#
#
#


